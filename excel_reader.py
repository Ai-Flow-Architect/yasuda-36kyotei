"""
36協定 Excel読み取りモジュール
社労士事務所向け36協定Excelテンプレート（43列: A〜AQ）からデータを読み取る
"""
import logging
from typing import Any

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path

# ロガー設定
logger = logging.getLogger(__name__)

# Excel列マッピング（A=1, B=2, ... AQ=43）
COLUMN_MAP: dict[str, int] = {
    "更新月": 1,                    # A
    "起算日_月": 2,                  # B
    "起算日_年": 3,                  # C
    "事業所名": 4,                   # D
    "事業主名": 5,                   # E
    "電話番号": 6,                   # F
    "事業の種類": 7,                 # G
    "時間外_事由": 8,                # H
    "時間外_業務の種類": 9,          # I
    "労働者数": 10,                  # J
    "18歳未満_労働者数": 11,         # K
    "所定労働時間": 12,              # L
    "延長時間_1日": 13,              # M
    "延長時間_1ヶ月": 14,            # N
    "時間外_期間": 15,               # O
    "休日_事由": 16,                 # P
    "休日_業務の種類": 17,           # Q
    "所定休日": 18,                  # R
    "休日労働_日数": 19,             # S
    "始業終業時刻": 20,              # T
    "休日_期間": 21,                 # U
    "特別条項の有無": 22,            # V
    "特別_理由": 23,                 # W
    "特別_業務の種類": 24,           # X
    "特別_労働者数": 25,             # Y
    "特別_延長時間": 26,             # Z
    "特別_超過回数": 27,             # AA
    "特別_延長時間_月": 28,          # AB
    "特別_割増賃金率": 29,           # AC
    "特別_延長時間_年": 30,          # AD
    "特別_手続き": 31,               # AE
    "特別_健康措置_番号": 32,        # AF
    "特別_健康措置_内容": 33,        # AG
    "誓約チェック": 34,              # AH
    "労働者代表_職": 35,             # AI
    "労働者代表_氏名": 36,           # AJ
    "過半数労働者_チェック": 37,     # AK
    "協定締結日": 38,                # AL
    "届出作成日": 39,                # AM
    "事業主職名": 40,                # AN
    "届出_事業主名": 41,             # AO
    "所轄労働局": 42,                # AP
    "所轄労基署": 43,                # AQ
}

# メール宛名の担当者名列を検出するためのヘッダーキーワード
# 運用シート（回収シート）は担当者名の列位置が固定でない（I列等）ため、
# 列番号ではなくヘッダー文字列で検出する（バグ②: 宛名が事業所コードになる問題対応）
RECIPIENT_HEADER_KEYWORDS: tuple[str, ...] = (
    "担当者名", "ご担当者", "宛名", "宛先担当者", "担当者", "ご担当",
)

# メールアドレス列（デモ用に追加）
EMAIL_COLUMN: int = 44  # AR列

# 様式パターン手動上書き列（AS列: 10/10_2 など自動判定外を明示指定するときに使用）
FORM_PATTERN_OVERRIDE_COLUMN: int = 45  # AS列

# 事業所番号列（AT列: "0001"など。ファイル名先頭番号とのマッチングに使用。空欄でも可）
OFFICE_NUMBER_COLUMN: int = 46  # AT列


# ---------- 様式判定ルール（優先順位付きリスト構造） ----------
# 各ルールは (判定関数, 返却値) のタプル。上から順に評価し、最初にTrueを返したルールが適用される。

def _is_yuuyo_minashi(record: dict[str, str]) -> bool:
    """適用猶予 + 事業場外みなし労働"""
    事業種類 = record.get("事業の種類", "")
    業務種類 = record.get("時間外_業務の種類", "")
    猶予キーワード = ["運転", "建設", "医師", "鹿児島", "沖縄", "砂糖"]
    if not any(kw in 事業種類 or kw in 業務種類 for kw in 猶予キーワード):
        return False
    return "みなし" in record.get("特別_手続き", "") or "事業場外" in 業務種類


def _is_yuuyo(record: dict[str, str]) -> bool:
    """適用猶予事業（運転・建設・医師等）"""
    事業種類 = record.get("事業の種類", "")
    業務種類 = record.get("時間外_業務の種類", "")
    猶予キーワード = ["運転", "建設", "医師", "鹿児島", "沖縄", "砂糖"]
    return any(kw in 事業種類 or kw in 業務種類 for kw in 猶予キーワード)


def _is_kenkyuu(record: dict[str, str]) -> bool:
    """研究開発業務"""
    業務種類 = record.get("時間外_業務の種類", "")
    研究開発キーワード = ["研究", "新技術", "新商品"]
    return any(kw in 業務種類 for kw in 研究開発キーワード)


def _is_tokubetsu(record: dict[str, str]) -> bool:
    """特別条項付き"""
    特別条項 = record.get("特別条項の有無", "")
    if 特別条項 and 特別条項 not in ("□", "なし", "無", ""):
        return True
    if record.get("特別_理由", "") or record.get("特別_延長時間_月", ""):
        return True
    return False


# 優先順位付きルールリスト: (判定関数, 様式コード)
# 上から順に評価し、最初にマッチしたルールの様式コードを返す
FORM_TYPE_RULES: list[tuple[Any, str]] = [
    (_is_yuuyo_minashi, "9_5"),  # 適用猶予＋事業場外みなし
    (_is_yuuyo,         "9_4"),  # 適用猶予事業
    (_is_kenkyuu,       "9_3"),  # 研究開発業務
    (_is_tokubetsu,     "9_2"),  # 特別条項付き
]

# デフォルト様式コード（どのルールにもマッチしない場合）
DEFAULT_FORM_TYPE: str = "9"


def _detect_recipient_column(ws: Worksheet) -> int | None:
    """ヘッダー行(1行目)から「担当者名／宛名」列の列番号を検出する。

    運用シート（回収シート）は担当者名列の位置がテンプレートと異なる（I列など）。
    列番号固定だと事業主名(E列)等を誤って宛名に使い、結果として
    事業所コードが宛名に表示される（バグ②）。ヘッダー文字列で検出して
    位置非依存にする。完全一致を優先し、なければ部分一致でフォールバック。
    """
    headers: list[tuple[int, str]] = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is not None:
            headers.append((col, str(v).strip()))

    # 1. 完全一致（最も信頼度が高い）
    for kw in RECIPIENT_HEADER_KEYWORDS:
        for col, h in headers:
            if h == kw:
                return col
    # 2. 部分一致（「担当者名（ご署名者）」等の表記ゆれを吸収）
    for kw in RECIPIENT_HEADER_KEYWORDS:
        for col, h in headers:
            if kw in h:
                return col
    return None


def _looks_like_office_code(value: str, office_number: str = "") -> bool:
    """値が事業所コード（宛名に使ってはいけない値）かどうか判定する。

    バグ②: 担当者名のつもりの列に事業所コードが入っていた場合、
    "0001 様" のような不自然な宛名を絶対に出さないためのガード。
    """
    v = str(value).strip()
    if not v:
        return True
    # 数字・記号のみ（例: "0001", "12-3", "001号"）は人名ではない
    if not any(ch.isalpha() or "぀" <= ch <= "ヿ" or "一" <= ch <= "鿿" for ch in v):
        return True
    # 事業所番号と一致（ゼロ埋め揺れも吸収）
    on = str(office_number).strip()
    if on and (v == on or v.zfill(4) == on.zfill(4)):
        return True
    return False


def read_excel(file_path: str) -> tuple[list[dict[str, str]], list[str]]:
    """Excelファイルから全行のデータを読み取る

    Args:
        file_path: Excelファイルのパス

    Returns:
        (レコード辞書のリスト, 警告メッセージのリスト)
    """
    logger.info("Excel読み取り開始: %s", file_path)
    wb: Workbook = openpyxl.load_workbook(file_path, data_only=True)
    ws: Worksheet = wb.active

    records: list[dict[str, str]] = []
    warnings: list[str] = []

    # メール宛名に使う担当者名列をヘッダーから検出（位置非依存・バグ②対応）
    recipient_col = _detect_recipient_column(ws)
    if recipient_col:
        logger.info("担当者名列を検出: 列%d", recipient_col)

    # 2行目からデータ行（1行目はヘッダー）
    for row_num in range(2, ws.max_row + 1):
        # 事業所名（D列）が空ならスキップ
        if not ws.cell(row=row_num, column=4).value:
            logger.debug("行%d: 事業所名が空のためスキップ", row_num)
            continue

        record: dict[str, str] = {}
        for key, col in COLUMN_MAP.items():
            cell_value = ws.cell(row=row_num, column=col).value
            record[key] = str(cell_value).strip() if cell_value is not None else ""

        # メールアドレス（AR列）、なければF列（6列目）のフォールバック
        # F列は電話番号列として定義されているため、@を含む場合のみメールとして扱う
        email_val = ws.cell(row=row_num, column=EMAIL_COLUMN).value
        email_str = str(email_val).strip() if email_val else ""
        if not email_str and ws.max_column >= 6:
            f_val = ws.cell(row=row_num, column=6).value
            f_str = str(f_val).strip() if f_val else ""
            if "@" in f_str:
                email_str = f_str
        record["メールアドレス"] = email_str

        # 案内文（I列=9列目: 回収シートのメール本文）
        案内_val = ws.cell(row=row_num, column=9).value
        record["案内文"] = str(案内_val).strip() if 案内_val else ""

        # 事業所番号（AT列）: "0001"など。空欄の場合は行番号を4桁ゼロ埋めで自動生成
        office_num_val = ws.cell(row=row_num, column=OFFICE_NUMBER_COLUMN).value
        office_num_str = str(office_num_val).strip() if office_num_val else ""
        if not office_num_str:
            office_num_str = str(row_num - 1).zfill(4)  # 2行目→0001, 3行目→0002...
        record["事業所番号"] = office_num_str

        # メール宛名用 担当者名（バグ②）
        # 優先: ヘッダー検出した担当者名列 → 事業主名(E列) → 空
        # 事業所コードが混入していたら宛名に使わない（ガード）
        担当者名 = ""
        if recipient_col:
            rv = ws.cell(row=row_num, column=recipient_col).value
            rv_str = str(rv).strip() if rv is not None else ""
            if rv_str and not _looks_like_office_code(rv_str, office_num_str):
                担当者名 = rv_str
        if not 担当者名:
            ej = record.get("事業主名", "")
            if ej and not _looks_like_office_code(ej, office_num_str):
                担当者名 = ej
        record["担当者名"] = 担当者名

        # 様式パターン手動上書き（AS列）: 10/10_2 など自動判定外を明示指定
        override_val = ws.cell(row=row_num, column=FORM_PATTERN_OVERRIDE_COLUMN).value
        override_str = str(override_val).strip() if override_val else ""

        # 入力バリデーション
        row_warnings: list[str] = validate_record(record, row_num)
        if row_warnings:
            warnings.extend(row_warnings)

        # 様式パターンを自動判定（上書き列が入力されていれば優先）
        if override_str:
            record["様式パターン"] = override_str
            logger.debug("行%d: %s → 様式%s (手動上書き)", row_num, record["事業所名"], override_str)
        else:
            record["様式パターン"] = detect_form_type(record)
            logger.debug("行%d: %s → 様式%s", row_num, record["事業所名"], record["様式パターン"])
        logger.debug("行%d: %s → 様式%s", row_num, record["事業所名"], record["様式パターン"])

        records.append(record)

    wb.close()

    # バリデーション警告を表示
    if warnings:
        logger.warning("入力データの警告 (%d件)", len(warnings))
        for w in warnings:
            logger.warning("  %s", w)

    logger.info("Excel読み取り完了: %d件", len(records))
    return records, warnings


def validate_record(record: dict[str, str], row_num: int) -> list[str]:
    """レコードのバリデーション（警告レベル）

    Args:
        record: 検証対象のレコード辞書
        row_num: Excelの行番号

    Returns:
        警告メッセージのリスト
    """
    warnings: list[str] = []
    事業所名: str = record.get("事業所名", "不明")

    # 必須項目チェック
    required: list[str] = ["事業主名", "事業の種類", "時間外_業務の種類", "労働者数"]
    for field in required:
        if not record.get(field, ""):
            warnings.append(f"行{row_num} [{事業所名}]: 「{field}」が未入力です")

    # 数値チェック
    numeric_fields: list[str] = ["労働者数", "延長時間_1日", "延長時間_1ヶ月"]
    for field in numeric_fields:
        val: str = record.get(field, "")
        if val and not val.replace(".", "").isdigit():
            warnings.append(f"行{row_num} [{事業所名}]: 「{field}」が数値ではありません: {val}")

    # 延長時間の上限チェック（月45時間を超える場合は特別条項が必要）
    try:
        月時間: float = float(record.get("延長時間_1ヶ月", "0") or "0")
        特別条項: str = record.get("特別条項の有無", "")
        if 月時間 > 45 and 特別条項 in ("□", "なし", "無", ""):
            warnings.append(f"行{row_num} [{事業所名}]: 月{月時間}時間は45時間超ですが特別条項が未設定です")
    except ValueError:
        pass

    return warnings


def detect_form_type(record: dict[str, str]) -> str:
    """Excelデータから36協定届の様式パターンを自動判定する

    優先順位付きルールリスト（FORM_TYPE_RULES）を上から順に評価し、
    最初にマッチしたルールの様式コードを返す。
    どのルールにもマッチしない場合はデフォルト（"9"）を返す。

    Args:
        record: Excelから読み取ったレコード辞書

    Returns:
        "9"   : 様式第9号（一般条項・月45h以内）
        "9_2" : 様式第9号の2（特別条項付き）
        "9_3" : 様式第9号の3（研究開発業務）
        "9_4" : 様式第9号の4（適用猶予事業）
        "9_5" : 様式第9号の5（適用猶予＋事業場外みなし）
    """
    for rule_func, form_type in FORM_TYPE_RULES:
        if rule_func(record):
            logger.debug("様式判定: ルール '%s' にマッチ → %s", rule_func.__name__, form_type)
            return form_type

    logger.debug("様式判定: デフォルト → %s", DEFAULT_FORM_TYPE)
    return DEFAULT_FORM_TYPE


if __name__ == "__main__":
    # テスト用
    import sys
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    if len(sys.argv) > 1:
        records = read_excel(sys.argv[1])
        for r in records:
            print(f"事業所: {r['事業所名']} | 様式: {r['様式パターン']} | メール: {r['メールアドレス']}")
    else:
        print("使い方: python excel_reader.py <Excelファイルパス>")
