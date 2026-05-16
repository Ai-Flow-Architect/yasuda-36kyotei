"""
クライアント報告バグ 2件の回帰テスト（2026-05-16）

バグ①: 1事業所に協定書が複数あると添付できない
バグ②: メール本文の宛名が事業所コードに変換される
"""
import tempfile
from pathlib import Path

import openpyxl

from excel_reader import read_excel, _detect_recipient_column, _looks_like_office_code
from word_matcher import match_word_files_multi, build_match_table
from mail_sender import build_email_body


# ============================================================
# バグ①: 複数協定書マッチング
# ============================================================
def _touch(p: Path):
    p.write_bytes(b"dummy")
    return p


def test_bug1_multi_kyotei_by_office_number():
    """同一事業所番号プレフィックスの複数協定書を全件返す"""
    with tempfile.TemporaryDirectory() as d:
        dp = Path(d)
        f1 = _touch(dp / "0001_36協定書_株式会社サンプル商事_様式第9号.docx")
        f2 = _touch(dp / "0001_36協定書_株式会社サンプル商事_様式第9号の2.docx")
        f3 = _touch(dp / "0002_36協定書_別会社_様式第9号.docx")
        got = match_word_files_multi("株式会社サンプル商事", [f1, f2, f3], office_number="0001")
        assert set(got) == {f1, f2}, got
        assert f3 not in got


def test_bug1_multi_kyotei_by_company_name():
    """事業所名一致でも複数様式を全件返す"""
    with tempfile.TemporaryDirectory() as d:
        dp = Path(d)
        f1 = _touch(dp / "36協定書_テスト工業_様式第9号.docx")
        f2 = _touch(dp / "36協定書_テスト工業_様式第9号の2.docx")
        got = match_word_files_multi("テスト工業", [f1, f2], office_number="")
        assert set(got) == {f1, f2}, got


def test_bug1_single_still_works():
    """単一協定書も従来どおり1件返す（後方互換）"""
    with tempfile.TemporaryDirectory() as d:
        dp = Path(d)
        f1 = _touch(dp / "0001_36協定書_単独社.docx")
        got = match_word_files_multi("単独社", [f1], office_number="0001")
        assert got == [f1]


def test_bug1_build_match_table_carries_all_paths():
    """build_match_table が _matched_paths に全件を載せる"""
    with tempfile.TemporaryDirectory() as d:
        dp = Path(d)
        f1 = _touch(dp / "0001_36協定書_A社_様式第9号.docx")
        f2 = _touch(dp / "0001_36協定書_A社_様式第9号の2.docx")
        records = [{"事業所名": "A社", "事業所番号": "0001", "メールアドレス": "a@example.invalid"}]
        table = build_match_table(records, [f1, f2])
        row = table[0]
        assert row["件数"] == 2
        assert set(row["_matched_paths"]) == {f1, f2}
        assert row["_matched_path"] == row["_matched_paths"][0]  # 後方互換


# ============================================================
# バグ②: 宛名が事業所コードになる
# ============================================================
def _make_excel(path: str, *, e_value: str, recipient_header: str | None, recipient_value: str):
    """回収シート風Excelを作る。E列=事業主名位置、任意列に担当者名ヘッダー"""
    wb = openpyxl.Workbook()
    ws = wb.active
    # ヘッダー
    ws.cell(row=1, column=4, value="事業所名")
    ws.cell(row=1, column=5, value="事業主名")
    if recipient_header:
        ws.cell(row=1, column=9, value=recipient_header)  # I列
    ws.cell(row=1, column=44, value="送信先メールアドレス")
    # データ
    ws.cell(row=2, column=4, value="株式会社サンプル商事")
    ws.cell(row=2, column=5, value=e_value)
    if recipient_header:
        ws.cell(row=2, column=9, value=recipient_value)
    ws.cell(row=2, column=44, value="to@example.invalid")
    wb.save(path)


def test_bug2_recipient_column_detected_by_header():
    """I列ヘッダー『担当者名』を検出し宛名に使う"""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as t:
        path = t.name
    _make_excel(path, e_value="0001", recipient_header="担当者名", recipient_value="山田太郎")
    records, _ = read_excel(path)
    assert records[0]["担当者名"] == "山田太郎"
    body = build_email_body(records[0], {"担当者名": "担当 花子"})
    assert "山田太郎 様" in body
    assert "0001 様" not in body


def test_bug2_office_code_in_E_does_not_become_atena():
    """担当者名列が無く事業主名がコードでも宛名にコードを出さない"""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as t:
        path = t.name
    _make_excel(path, e_value="0001", recipient_header=None, recipient_value="")
    records, _ = read_excel(path)
    body = build_email_body(records[0], {"担当者名": "担当 花子"})
    assert "0001 様" not in body
    assert "ご担当者 様" in body


def test_bug2_normal_person_name_in_E_still_used():
    """事業主名が人名なら従来どおり宛名に使う（後方互換）"""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as t:
        path = t.name
    _make_excel(path, e_value="田中一郎", recipient_header=None, recipient_value="")
    records, _ = read_excel(path)
    body = build_email_body(records[0], {"担当者名": "担当 花子"})
    assert "田中一郎 様" in body


def test_bug2_guard_helpers():
    """コード判定ヘルパーの境界"""
    assert _looks_like_office_code("0001")
    assert _looks_like_office_code("12-3")
    assert _looks_like_office_code("", "0001")
    assert _looks_like_office_code("0001", "1")  # zfill一致
    assert not _looks_like_office_code("田中一郎")
    assert not _looks_like_office_code("山田 太郎")


def test_bug2_build_email_body_direct_guard():
    """build_email_body 単体でもコードガードが効く（二重ガード）"""
    body = build_email_body(
        {"事業所名": "X社", "担当者名": "0007", "事業所番号": "0007"},
        {"担当者名": "担当 花子"},
    )
    assert "0007 様" not in body
    assert "ご担当者 様" in body


if __name__ == "__main__":
    import sys
    import pytest
    sys.exit(pytest.main([__file__, "-q"]))
